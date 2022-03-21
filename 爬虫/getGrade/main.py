import time
import urllib.parse
from http.client import HTTPResponse
from bs4 import BeautifulSoup #网页解析
import urllib.request as request,urllib.error as urlerror #制定url，获取网页数据
import xlrd,xlwt #进行excel操作


import hashlib
from student import Student


def get_rep(stu):
    """获取会话session和加密字符串"""
    cookies = 'semester.id=140;'
    url = 'http://jwgl.aust.edu.cn/eams/login.action'
    req = request.Request(url, data=stu.data, headers=stu.headers)
    rep = request.urlopen(req)  # type:HTTPResponse
    rep_head=rep.getheaders()
    for each in rep_head:
        if each[0]=='Set-Cookie':
            cookies+=each[1][0:each[1].find(';')+1]
    stu.set_cookie(cookies[0:-2])
    # time.sleep(0.5)
    return rep



def login(rep,stu):
    """实现用户登录"""
    url = 'http://jwgl.aust.edu.cn/eams/login.action'
    html=rep.read().decode()#type:HTTPResponse
    html=BeautifulSoup(html,'html.parser')
    js=html.select('script[type="text/javascript"]')
    js=str(js.pop())
    # print(js)
    key=js[js.find("form['password'].value = CryptoJS.SHA1(")+len("form['password'].value = CryptoJS.SHA1('"):js.find("' + form['password'].value);")]

    # print(key)
    password=stu.password
    password=hashlib.sha1((key+password).encode('utf-8')).hexdigest()
    stu.data=bytes(urllib.parse.urlencode({'username':stu.id,
                                           'password':password,
                                           'session_locale':'zh_CN',
                                           'encodedPassword':''}),encoding='utf-8')
    req_login=request.Request(url,headers=stu.headers,data=stu.data)
    rep=request.urlopen(req_login)
    rep_html=rep.read().decode('utf-8')
    bs = BeautifulSoup(rep_html, 'html.parser')
    try:
        a=bs.select("a[href='/eams/security/my.action']")
        name=a[0].text
        stu.name=name[0:name.find('(')]
    except:
        a=bs.select("div[class='ui-state-error ui-corner-all'] > span")
        print(a[1].text)
    return rep_html

def get_grade_html(stu):
    """获取成绩单"""
    url = 'http://jwgl.aust.edu.cn/eams/teach/grade/course/person!search.action?semesterId=140&projectType='
    req = urllib.request.Request(url, headers=stu.headers)
    rep = urllib.request.urlopen(req)  # type:HTTPResponse
    return rep.read().decode('utf-8')

def data_soup(html):
    """解析成绩单数据"""
    data={}
    soup=BeautifulSoup(html,'html.parser')
    tr=soup.select('tbody > tr')
    for tds in tr:
        grad=[]
        for td in tds:
            try:
                text=td.text.strip().replace('\t','').replace('\n','')
                if text !='':
                    grad.append(text)
            except:
                pass
        data[grad[3]]=grad[6]
    return data


import openpyxl
from openpyxl.styles import PatternFill
def get_student_info():
    students=openpyxl.load_workbook("E:\Desktop\Study\zr.xlsx")
    sheetnames = students.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    stus = students.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容
    password={}
    # 获取sheet的最大行数和列数
    rows = stus.max_row
    cols = stus.max_column

    for r in range(2, rows+1):
        # if stus.cell(r,5).value=='':
        password[stus.cell(r, 6).value]=stus.cell(r, 2).value[-6:]
    print(rows)
    return password



def makexls():
    # 初始化数据
    pas=get_student_info()
    write = openpyxl.Workbook()  # 打开一个将写的文件
    w = write.create_sheet(index=0)  # 在将写的文件创建sheet
    fill=PatternFill("solid", fgColor="ff7575")


    row=2
    #列名
    gradeType= {}
    gradeType_index=2

    for k,v in pas.items():
        stu = Student(str(k),str(v))
        rep=get_rep(stu)#type:HTTPResponse
        try:
            rep=login(rep,stu)
        except:
            print('id',k,'登录失败')
            break
        html=get_grade_html(stu)
        try:
            data=data_soup(html)
            w.cell(row,1).value=stu.name
            for k, v in data.items():#k:gradetype  v:grade
                if k not in gradeType.values():
                    gradeType[gradeType_index]=k
                    w.cell(1,gradeType_index).value=k
                    gradeType_index+=1
                for index,g in gradeType.items():
                    if g==k:
                        col=index
                w.cell(row, col).value = v
                if len(v)>3:
                    w.cell(row, col).fill = fill
                elif int(v)<60 and v!=None:
                    w.cell(row, col).fill=fill
            print(stu.name, '获取成功')
            row += 1
        except:
            print(stu.name,'获取分数失败')
    print(gradeType)
    saveExcel = "全校成绩.xlsx"
    write.save(saveExcel)  # 一定要记得保存


def getone():
    stu = Student('2020304643','25371')
    rep = get_rep(stu)  # type:HTTPResponse
    try:
        html = login(rep, stu)
    except:
        pass
    html = get_grade_html(stu)
    data = data_soup(html)
    print(data)
    print(stu.name)



if __name__ == '__main__':
    # makexls()
    getone()

