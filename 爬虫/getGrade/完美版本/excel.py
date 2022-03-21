import openpyxl
from openpyxl.styles import PatternFill
from student import Student

def get_student_info():
    students=openpyxl.load_workbook("zr.xlsx")
    sheetnames = students.sheetnames  # 获取读文件中所有的sheet，通过名字的方式
    stus = students.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容

    allstu = []
    password={}
    # 获取sheet的最大行数和列数
    rows = stus.max_row
    cols = stus.max_column
    n=0
    for r in range(2, rows):
        n+=1
        if n<5000:#stus.cell(r,5).value=='物联网工程2001班':
            stu=Student(stus.cell(r, 6).value,stus.cell(r, 2).value[-6:])
            stu.classname=stus.cell(r,5).value
            allstu.append(stu)
    print(rows)
    return allstu



def write_student_info(sheet,data,name):
    try:
        fill = PatternFill("solid", fgColor="ff7575")
        row = sheet.row
        w=sheet.worksheet
        # 列名
        gradeType = sheet.gradetype
        gradeType_index = 2
        w.cell(row, 1).value = name
        for k, v in data.items():  # k:gradetype  v:grade
            if k not in gradeType.values():
                gradeType[gradeType_index] = k
                w.cell(1, gradeType_index).value = k
                gradeType_index += 1
            for index, g in gradeType.items():
                if g == k:
                    col = index
            w.cell(row, col).value = v
            if len(v) > 5:
                w.cell(row, col).fill = fill
            elif float(v) < 60 and v != None:
                w.cell(row, col).fill = fill
        sheet.row += 1
    except Exception as e:
        print(e)